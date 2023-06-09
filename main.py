#import a module to read file
import openpyxl

# choose a function (load_workbook)to load file to read and save as a variable
inv_file = openpyxl.load_workbook("inventory.xlsx")

# capture the sheet to work on and assign it to a variable
product_list = inv_file["Sheet1"]

# calculate how many product we have per supplier using a dictionary where a name of the company is key: product count is the value
products_per_supplier = {}

# create a dictionary for total product per supplier
total_value_per_supplier = {}

# creating dictn for printing product less than 10
products_under_10_inv = {}

#N.B: Go through each  row in sheet and get supplier name using for loops as many time of no products.looking up to (max_row) syntax in documentn of module

#use range to iterate through the rows and to start with row 2 where d datas to use are,by default it start from 0,to capture d data in last row,add 1 to range

for product_row in range(2, product_list.max_row + 1):

# get the first info needed in the sheet which is supplier name in fourth column through row n column cell,using dot value helps show the real content in the cell

    supplier_name = product_list.cell(product_row, 4).value

# to calculate total number product item and price in each cell(row n column)
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

# get  product number
    product_num = product_list.cell(product_row, 1).value

# to write all calcln. into the excel sheet column next to column with data
    inventory_price = product_list.cell(product_row, 5)

#CALCULATION NUMBER OF PRODUCT PER SUPLIER.

#condition to check incremental count of suppliers if true then execute next code with add up to supplier count.

    if supplier_name in products_per_supplier:

# build each supplier and her product quantity with dictionary and aslo show current product of supplier/alternatively we can invoke .get(supplier_name)method instead[supplier_name]

        current_num_products =  products_per_supplier[supplier_name]

# checks for the key value in dict.during iteration if same key(Supplier) is noted it stores an increment for the value  +1 and assign it back as a loop into the dictionary

        products_per_supplier[supplier_name] = current_num_products +1

#if the condition for incrementing value to a supplier differs ,it mean a new supplier is detected,so it record new supplier and continue loop else it increment value for supplier
    else:
        products_per_supplier[supplier_name] = 1


# CALCULATION OF TOTAL VALUE OF INVENTORY PER SUPPLIER

# considering the iteration for the total of first supplier.
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] =  current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# LOGIC PRODUCT WITH INVENTORY LESS THAN 10
    if  inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

# ADD VALUE FOR TOTAL INVENTORY PRICE .set value for the cells
        inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# calling a function explicitly to save the file differently from the initial sheet worked on and store all calcn in colm 5,name of file in bracket below as a new sheet.
inv_file.save("inventory_with_total_value.xlsx")

